import requests
from django import forms
from django.contrib import admin

import csv
from django.http import HttpResponse
import io

from xariduzex.models import XariduzexCheck
from .models import *

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


class CategoryAdmin(admin.ModelAdmin):
    list_display = ('category_id', 'name',)
    search_fields = ('category_id', 'name',)
    list_filter = ('category_id', 'name',)
    ordering = ('category_id', 'name',)
    fieldsets = (
        (None, {
            'fields': ('category_id', "name")
        }),
    )

    readonly_fields = ("name",)
    actions = [
        'export_as_csv',
    ]

    @staticmethod
    def get_all_categories(id=None):
        url = requests.get(f'https://xarid-api-trade.uzex.uz/Lib/GetCategories')
        data = url.json()
        if id is not None:
            data = [category for category in data if category['id'] == id]
        return data

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['category_id'] = forms.ChoiceField(
            choices=[(i['id'], f"{i['id']}-{i['name']}") for i in self.get_all_categories(id=request.GET.get('id'))])
        return form

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        if not change:
            for i in self.get_all_categories():
                if i['id'] == int(form.cleaned_data['category_id']):
                    obj.name = i['name']
                    data = requests.get(f'https://xarid-api-trade.uzex.uz/Lib/GetProducts/{i["id"]}')
                    for j in data.json():
                        Product.objects.create(
                            product_id=j['id'],
                            name=j['name'],
                            category=obj,
                            product_code=j['product_code'])
                    CheckedTender.objects.filter(category_id=i['id']).delete()
                    break
        obj.save()

    def has_change_permission(self, request, obj=None):
        return False

    def export_as_csv(self, request, queryset):
        f = io.StringIO()
        writer = csv.writer(f)
        writer.writerow(["name"])
        for s in queryset:
            writer.writerow([s.name])
        f.seek(0)
        response = HttpResponse(f, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=category.csv'
        return response

    export_as_csv.short_description = "Export Selected as CSV"


admin.site.register(Category, CategoryAdmin)


class ProductAdmin(admin.ModelAdmin):
    list_display = ('product_id', 'name', 'category', 'product_code')
    search_fields = ('product_id', 'name', 'category__name', 'product_code')
    list_filter = ('category',)
    ordering = ('product_id', 'name')
    fieldsets = (
        (None, {
            'fields': ('product_id', 'name', 'category', 'product_code')
        }),
    )

    actions = [
        'export_as_csv',
    ]

    @staticmethod
    def export_as_csv(request, queryset):
        f = io.StringIO()
        writer = csv.writer(f)
        writer.writerow(["product_id", "name", "category", "product_code"])
        for s in queryset:
            writer.writerow([s.product_id, s.name, s.category, s.product_code])
        f.seek(0)
        response = HttpResponse(f, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=product.csv'
        return response

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def get_readonly_fields(self, request, obj=None):
        return [field.name for field in obj._meta.fields]


admin.site.register(Product, ProductAdmin)


class TenderAdmin(admin.ModelAdmin):
    list_display = ('tender_id', 'name', 'display_number', 'type', 'start_date', 'end_date', 'cost', 'currency',
                    'seller_name', 'seller_tin', 'region_name', 'district_name', 'seller_id', 'category', 'url')
    search_fields = ('tender_id', 'name', 'display_number', 'type', 'start_date', 'end_date', 'cost', 'currency',
                     'seller_name', 'seller_tin', 'region_name', 'district_name', 'seller_id', 'category__name', 'url')
    list_filter = ('type', 'currency', 'category')
    ordering = ('tender_id', 'name', 'display_number', 'type', 'start_date', 'end_date', 'cost', 'currency',
                'seller_name', 'seller_tin', 'region_name', 'district_name', 'seller_id', 'category', 'url')
    fieldsets = (
        (None, {
            'fields': ('tender_id', 'name', 'display_number', 'type', 'start_date', 'end_date', 'cost', 'currency',
                       'seller_name', 'seller_tin', 'region_name', 'district_name', 'seller_id', 'category', 'url')
        }),
    )

    actions = [
        'export_as_csv',
        'export_to_excel',
        'export_to_json',
    ]

    def export_as_csv(self, request, queryset, *args, **kwargs):
        f = io.StringIO()
        writer = csv.writer(f)
        writer.writerow(
            ["tender_id", "name", "display_number", "type", "start_date", "end_date", "cost", "currency",
             "seller_name", "seller_tin", "region_name", "district_name", "seller_id", "category", "url"])
        for s in queryset:
            writer.writerow(
                [s.tender_id, s.name, s.display_number, s.type, s.start_date, s.end_date, s.cost, s.currency,
                 s.seller_name, s.seller_tin, s.region_name, s.district_name, s.seller_id, s.category, s.url])
        f.seek(0)
        response = HttpResponse(f, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=tender.csv'
        return response

    def export_to_json(self, request, queryset):
        data = []
        for s in queryset:
            data.append(
                {
                    "tender_id": s.tender_id,
                    "name": s.name,
                    "display_number": s.display_number,
                    "type": s.type,
                    "start_date": s.start_date,
                    "end_date": s.end_date,
                    "cost": s.cost,
                    "currency": s.currency,
                    "seller_name": s.seller_name,
                    "seller_tin": s.seller_tin,
                    "region_name": s.region_name,
                    "district_name": s.district_name,
                    "seller_id": s.seller_id,
                    "category": s.category,
                    "url": s.url
                }
            )
        response = HttpResponse(data, content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename=tender.json'
        return response

    def export_to_excel(self, request, queryset):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Tenders"

        columns = [
            "tender_id", "name", "display_number", "type", "start_date", "end_date", "cost", "currency",
            "seller_name", "seller_tin", "region_name", "district_name", "seller_id", "category", "url"
        ]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for tender in queryset:
            row_num += 1
            row = [
                tender.tender_id, tender.name, tender.display_number, "Tender" if tender.type == "1" else "Auction",
                str(tender.start_date), str(tender.end_date),
                tender.cost, tender.currency, tender.seller_name, tender.seller_tin, tender.region_name,
                tender.district_name,
                tender.seller_id, tender.category.name if tender.category else "",
                "https://etender.uzex.uz/lot/" + str(tender.tender_id)
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        for col_num in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].auto_size = True

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=tenders.xlsx'
        workbook.save(response)
        return response

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def get_readonly_fields(self, request, obj=None):
        return [field.name for field in obj._meta.fields]


admin.site.register(Tender, TenderAdmin)


class TenderProductAdmin(admin.ModelAdmin):
    list_display = ('tender', 'product')
    search_fields = ('tender__name', 'product__name')
    list_filter = ('tender', 'product')
    ordering = ('tender', 'product')
    fieldsets = (
        (None, {
            'fields': ('tender', 'product')
        }),
    )

    actions = [
        'export_as_csv',
    ]

    def export_as_csv(self, request, queryset):
        f = io.StringIO()
        writer = csv.writer(f)
        writer.writerow(["tender", "product"])
        for s in queryset:
            writer.writerow([s.tender, s.product])
        f.seek(0)
        response = HttpResponse(f, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=tender_product.csv'
        return response

    export_as_csv.short_description = "Export Selected as CSV"


admin.site.register(TenderProduct, TenderProductAdmin)


class CheckedTenderAdmin(admin.ModelAdmin):
    list_display = ('tender_id', 'category_id')
    search_fields = ('tender_id', 'category_id')
    list_filter = ('tender_id', 'category_id')
    ordering = ('tender_id', 'category_id')
    fieldsets = (
        (None, {
            'fields': ('tender_id', 'category_id')
        }),
    )

    actions = [
        'export_as_csv',
    ]

    def export_as_csv(self, request, queryset):
        f = io.StringIO()
        writer = csv.writer(f)
        writer.writerow(["tender_id", "category_id"])
        for s in queryset:
            writer.writerow([s.tender_id, s.category_id])
        f.seek(0)
        response = HttpResponse(f, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=checked_tender.csv'
        return response

    export_as_csv.short_description = "Export Selected as CSV"


admin.site.register(CheckedTender, CheckedTenderAdmin)


# XariduzexCheck
class XariduzexCheckAdmin(admin.ModelAdmin):
    list_display = ('tender_id', 'category')
    search_fields = ('tender_id', 'category')
    list_filter = ('tender_id', 'category')
    ordering = ('tender_id', 'category')
    fieldsets = (
        (None, {
            'fields': ('tender_id', 'category')
        }),
    )

    actions = [
        'export_as_csv',
    ]

    def export_as_csv(self, request, queryset):
        f = io.StringIO()
        writer = csv.writer(f)
        writer.writerow(["tender_id", "category"])
        for s in queryset:
            writer.writerow([s.tender_id, s.category])
        f.seek(0)
        response = HttpResponse(f, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=xariduzex_check.csv'
        return response

    export_as_csv.short_description = "Export Selected as CSV"


admin.site.register(XariduzexCheck, XariduzexCheckAdmin)